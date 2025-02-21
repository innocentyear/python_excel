from openpyxl import Workbook, load_workbook
wb = load_workbook("./test.xlsx")
ws = wb.active
ws.title = "工作表1"
print(ws.title)



print(wb.sheetnames)


#wb.save("./test.xlsx")