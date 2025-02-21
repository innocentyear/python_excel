from openpyxl import Workbook, load_workbook
wb = load_workbook("./test.xlsx")
ws1 = wb.active
ws1.title = "Sheet1"
#print(ws.title)

ws2 = wb.create_sheet("Sheet2",1) #创建工作表2，插入到第二个位置
ws3 = wb.create_sheet("Sheet3",2)#创建工作表3，插入到第三个位置

#ws =wb["Sheet3"] #通过工作表名获取工作表
#print(ws.title)

wb.move_sheet(ws1,-2) #将工作表向前移动2位

del wb["Sheet3"] #删除工作表3


cp_sheet = wb.copy_worksheet(ws1) #复制工作表1
print(cp_sheet.title)   


print(wb.sheetnames)


#wb.save("./test.xlsx")